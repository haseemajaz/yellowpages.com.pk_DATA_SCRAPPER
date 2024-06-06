import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import filedialog
import threading
from openpyxl import Workbook, load_workbook

def scrape_from_urls_file():
    # Open file dialog to select the text file containing URLs
    file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])

    if file_path:
        with open(file_path, 'r') as file:
            urls = file.readlines()
        
        if urls:
            # Create a new thread to perform scraping
            scrape_thread = threading.Thread(target=scrape_urls, args=(urls,))
            scrape_thread.start()
        else:
            error_label.config(text="No URLs found in the selected file.")

def scrape_urls(urls):
    # Create a new GUI window
    window = tk.Toplevel(root)
    window.title("Scraped Content")

    # Create a text widget to display the scraped content
    text_widget = tk.Text(window, wrap=tk.WORD)
    text_widget.pack(expand=True, fill=tk.BOTH)

    # Create or load an existing Excel workbook
    excel_file = "scraped_data.xlsx"
    try:
        wb = load_workbook(excel_file)
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active

    # Loop through the URLs and scrape each website for content
    for url_index, url in enumerate(urls, start=1):
        scrape_website(url.strip(), text_widget, ws, url_index)
        # Save the Excel file after scraping each URL
        wb.save(excel_file)

def scrape_website(url, text_widget, ws, url_index):
    # Send a GET request to the website
    response = requests.get(url)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Find all elements with class="listing-url"
        listing_urls = soup.find_all(class_='listing-url')
        
        # Find all elements with class="listing-title"
        listing_titles = soup.find_all(class_='listing-title')
        
        # Find all elements with class="listing-phone"
        listing_phones = soup.find_all(class_='listing-phone')
        
        # Extract the content of each listing-url, listing-title, and listing-phone element
        for listing_url, listing_title, listing_phone in zip(listing_urls, listing_titles, listing_phones):
            url_content = listing_url.get_text(strip=True)
            title_content = listing_title.get_text(strip=True)
            phone_content = listing_phone.get_text(strip=True)
            if url_content and title_content and phone_content:  # Check if content is not empty
                text_widget.insert(tk.END, f"URL: {url_content}, Title: {title_content}, Phone: {phone_content}\n")
                # Write the content to the Excel file in the first, second, and third columns
                ws.cell(row=url_index, column=1, value=url_content)
                ws.cell(row=url_index, column=2, value=title_content)
                ws.cell(row=url_index, column=3, value=phone_content)
                url_index += 1
    else:
        text_widget.insert(tk.END, f"Failed to retrieve {url}. Status code: {response.status_code}\n")

# Create a Tkinter window
root = tk.Tk()
root.title("Web Scraping")

# Create a button to import URLs from a text file
import_button = tk.Button(root, text="Import URLs from File", command=scrape_from_urls_file)
import_button.pack(pady=10)

# Error label
error_label = tk.Label(root, fg="red")
error_label.pack()

# Start the Tkinter event loop
root.mainloop()
